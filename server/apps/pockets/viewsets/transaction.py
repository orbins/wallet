from decimal import Decimal
import logging
from typing import Type, Union

from django.http import FileResponse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework import viewsets, serializers, pagination
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response

from ..constants import TransactionTypes
from ..filters import TransactionFilter
from ..models import Transaction, TransactionCategory
from ..models.querysets import TransactionQuerySet
from ..serializers import (
    BalanceSerializer,
    TransactionCreateSerializer,
    TransactionRetrieveSerializer,
    TransactionGlobalSerializer,
)
from ..services import TransactionFileHandler

logger = logging.getLogger('__name__')


class TransactionViewSet(viewsets.ModelViewSet):
    pagination_class = pagination.LimitOffsetPagination
    pagination_class.default_limit = 20
    permission_classes = (IsAuthenticated,)
    filterset_class = TransactionFilter

    def get_serializer_class(self) -> Type[serializers.ModelSerializer]:
        if self.action == 'total':
            serializer_class = TransactionGlobalSerializer
        elif self.action == 'get_balance':
            serializer_class = BalanceSerializer
        elif self.action in ('create', 'update', 'partial_update'):
            serializer_class = TransactionCreateSerializer
        else:
            serializer_class = TransactionRetrieveSerializer

        return serializer_class

    def get_queryset(self) -> TransactionQuerySet:
        queryset = Transaction.objects.filter(
            user=self.request.user,
        ).select_related('category',).order_by(
            '-transaction_date', '-id',
        )
        return queryset

    def get_object(self) -> Union[Transaction, dict[str, Decimal]]:
        queryset = self.get_queryset()
        if self.action == 'total':
            obj = self.filter_queryset(queryset).aggregate_totals()
        elif self.action == 'get_balance':
            obj = self.filter_queryset(queryset).calculate_balance()
        else:
            obj = super().get_object()

        return obj

    @action(methods=('GET',), detail=False, url_path='global')
    def total(self, request: Request, *args, **kwargs) -> Response:
        return super().retrieve(request, *args, **kwargs)

    @action(methods=('GET',), detail=False, url_path='balance')
    def get_balance(self, request: Request, *args, **kwargs) -> Response:
        return super().retrieve(request, *args, **kwargs)

    @action(methods=('GET',), detail=False, url_path='export')
    def export_data(self, request: Request, *args, **kwargs) -> FileResponse:
        transactions = self.get_queryset().only(
            'category', 'transaction_date',
            'amount', 'transaction_type'
        )
        file = TransactionFileHandler(
            transactions=transactions
        ).export_transactions_to_excel()

        return FileResponse(file, as_attachment=True, filename='transactions.xlsx')

    @action(methods=('POST',), detail=False, url_path='import')
    def import_data(self, request: Request, *args, **kwargs) -> Response:
        file = request.FILES['file']
        wb = load_workbook(file)
        ws = wb.active
        errors = []
        transactions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            transaction_date, transaction_type, category_name, amount = row
            if not transaction_date:
                continue
            if transaction_type != TransactionTypes.CHOICES_DICT[TransactionTypes.INCOME]:
                category, _ = TransactionCategory.objects.get_or_create(
                    user=request.user,
                    name__iexact=category_name,
                    defaults={'name': category_name}
                )
            else:
                category = None
            transaction = {
                'transaction_date': transaction_date.date(),
                'transaction_type': 'income' if transaction_type == 'Доход' else 'expense',
                'category': category.id if category else None,
                'amount': amount,
                'user': request.user
            }
            serializer = TransactionCreateSerializer(
                data=transaction,
                context={'request': request}
            )
            try:
                serializer.is_valid(raise_exception=True)
            except serializers.ValidationError as e:
                errors.append({
                    'transaction_row': row,
                    'errors': e.detail
                })
                continue
            transactions.append(transaction)
        if errors:
            return Response(data=errors, status=status.HTTP_400_BAD_REQUEST)
        serializer = TransactionCreateSerializer(
            data=transactions,
            context={'request': request},
            many=True
        )
        serializer.is_valid()
        serializer.save()
        return Response(data=None, status=status.HTTP_200_OK)






