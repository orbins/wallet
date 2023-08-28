from io import BytesIO

from openpyxl import Workbook, load_workbook
from rest_framework import serializers, status

from ..constants import TransactionTypes
from ..models import TransactionCategory
from ..serializers import TransactionCreateSerializer


class TransactionFileHandler:
    def __init__(self, transactions=None, file=None):
        """
        :param transactions: queryset модели Transaction
        """
        self.transactions = transactions or []
        try:
            self.wb = load_workbook(file) if file else Workbook()
            self.ws = self.wb.active
        except Exception:
            self.wb = None

    def save_transactions_to_worksheet(self):
        self.ws.append(
            ['Дата операции', 'Тип операции', 'Категория', 'Сумма']
        )
        for transaction in self.transactions:
            self.ws.append([
                transaction.transaction_date,
                'Доход' if transaction.transaction_type == 'income' else 'Расход',
                transaction.category.name if transaction.category else None,
                transaction.amount
            ])

    def save_file(self):
        file = BytesIO()
        self.wb.save(file)
        file.seek(0)

        return file

    def export_transactions_to_excel(self):
        self.save_transactions_to_worksheet()
        file = self.save_file()

        return file

    def load_data_from_worksheet(self, request):
        errors = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            transaction_date, transaction_type, category_name, amount = row
            if transaction_date == transaction_type:
                continue
            if transaction_type != TransactionTypes.CHOICES_DICT[TransactionTypes.INCOME]:
                category, _ = TransactionCategory.objects.get_or_create(
                    user=request.user,
                    name__iexact=category_name,
                    defaults={'name': category_name}
                )
            else:
                category = None
            transaction = {
                'transaction_date': transaction_date.date(),
                'transaction_type': 'income' if transaction_type == 'Доход' else 'expense',
                'category': category.id if category else None,
                'amount': amount,
                'user': request.user
            }
            serializer = TransactionCreateSerializer(
                data=transaction,
                context={'request': request}
            )
            try:
                serializer.is_valid(raise_exception=True)
            except serializers.ValidationError as e:
                errors.append({
                    'transaction_row': row,
                    'errors': e.detail
                })
                continue
            self.transactions.append(transaction)

        return errors, self.transactions

    def save_transactions(self, request):
        serializer = TransactionCreateSerializer(
            data=self.transactions,
            context={'request': request},
            many=True
        )
        serializer.is_valid()
        serializer.save()

    def import_transactions_from_excel(self, request):
        if not self.wb:
            return status.HTTP_400_BAD_REQUEST, {
                'error': "Ошибка чтения файла"
            }
        errors, self.transactions = self.load_data_from_worksheet(request)
        if errors:
            return status.HTTP_400_BAD_REQUEST, errors
        self.save_transactions(request)

        return status.HTTP_200_OK, None
